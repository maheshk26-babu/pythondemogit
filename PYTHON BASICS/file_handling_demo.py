import datetime

from openpyxl import Workbook, load_workbook



# fp=open("sample.txt", 'w')
# fp.write(str2)
fp=open("../sample.txt", 'r')
print(fp.readline(), end='')
for data in fp:
    print(data)

fp1 = open("img1.JPG", 'rb')# images we have to use rb--> read binary mode
f1 = open("Mypic.JPG", 'wb')
for i in fp1:
    print(fp1)
    f1.write(i)


file_contents=fp.read()
print("file contents are:",file_contents)
fp.close()
#
# wb= Workbook()
# ws= wb.active
# wb.create_sheet('Data')
# ws['A1']= 42
# ws.append([1,2,3])
# ws['A2']=datetime.datetime.now()
# ws.title='Title'
# wb.save('Sample.xlsx')
#
# wb2=load_workbook("../Sample.xlsx")
# sheet_names=wb2.sheetnames
# print(sheet_names)


