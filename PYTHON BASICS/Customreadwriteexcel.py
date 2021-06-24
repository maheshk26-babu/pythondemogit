# -*- coding: utf-8 -*-
# @File    : Customreadwriteexcel.py
# @Time    : 4/29/2021 3:53 PM
# @Author  : Mahesh
# @Email   : mahesh.babux.kandukuri@email

import openpyxl
from openpyxl import load_workbook


class ReadWriteExcel:
    wb = None
    excelfilepath = None

    def __init__(self, filepath):
        global wb
        global excelfilepath
        excelfilepath = filepath
        self.wb = openpyxl.load_workbook(excelfilepath)

    def readByColIndex(self, sheetname, rowindex, colindex):
        sheet = self.wb[sheetname]
        return sheet.cell(rowindex, colindex).value

    def readByColName(self, sheetname, rowindex, colname):
        sheet = self.wb[sheetname]
        colindex = 1
        while sheet.cell(row=1, column=colindex).value != '':
            if colname == sheet.cell(row=1, column=colindex).value:
                break
            colindex = colindex + 1
        return sheet.cell(rowindex, colindex).value

    def writeByColIndex(self, sheetname, rowindex, colindex, value):
        sheet = self.wb[sheetname]
        sheet.cell(rowindex, colindex).value = value
        self.wb.save(excelfilepath)

    def writeByColName(self, sheetname, rowindex, colname, value):
        sheet = self.wb[sheetname]
        colindex = 1
        while sheet.cell(row=1, column=colindex).value != '':
            if colname == sheet.cell(row=1, column=colindex).value:
                break
            colindex = colindex + 1
            sheet.cell(rowindex, colindex).value = value
            self.wb.save(excelfilepath)

    def getRowCount(self, sheetname):
        sheet = self.wb[sheetname]
        return sheet.max_row

    def getColCount(self, sheetname):
        sheet = self.wb[sheetname]
        return sheet.max_column
